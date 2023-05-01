import streamlit as st
import ping3
import openpyxl

def ping_ping3(hostname):
    timeout = 3
    response = ping3.ping(hostname, timeout=timeout)
    if response == None:
        result = "Fail"
    else:
        result = "Success"
    return result

 
filename = "test.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active
st.write('# Ping Try')
for x in range(1, ws.max_row + 1):
    hostname = ws.cell(row=x, column=2).value
    print(hostname, end="\n")
    ping_result = ping_ping3(hostname)
    print(ping_result)
    ws.cell(row=x, column=1).value = ping_result
    hostname
    ping_result

wb.save(filename) # 파일 저장
